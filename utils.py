################################################################################
#
# ARDD website general utils. 
# Created by Thomas Chia and Cindy Wu
# Medical Research by Sreya Devarakonda
# Created for the 2021 Congressional App Challenge
# Winning "webapp" of Virginia's 11th District
#
# Citations:
# The overlay image function was inspired by the OpenCV Documentation.
#   https://docs.opencv.org/trunk/d0/d86/tutorial_py_image_arithmetics.html
#
################################################################################

import numpy as np
import cv2 as cv2
import pandas as pd
from tensorflow.keras.preprocessing import image
from glaucoma_segmentation.utils import segment_image
from object_detection.utils import detect_image
import re
import base64
import openpyxl
from PIL import Image
from io import BytesIO

def convert_data(diagnosis_sheet, file_name):
    """
    Reads the data from the diagnosis sheet and transforms into a pandas dataframe.

    Parameters:
        diagnosis_sheet (str): NO USE
        file_name (str): The file name of the diagnosis sheet.
    
    Returns:
        Pandas dataframe.
    """
    data_ss = pd.read_excel('./uploads/sheets/' + 'summary_' + file_name[:-4] + '.xlsx')
    data_pd = {}
    for index, row in data_ss.iterrows():
        data_pd[index] = dict(row)
    return data_pd

def model_predict(img_path, file_name, yolo_model, disc_model, seg_model):
    """
    Creates the predicitons across object detection and image segmentation.

    Parameters:
        img_path: path to the saved image
        file_name: the name of the file to save INTO
        yolo_model: the Tensorflow model for the object detection model
        disc_model: the Tensorflow model for the disc detection model.
        seg_modle: the Tensorflow model for the segmentation model.
    
    Returns:
        im: the image
        diagnosis_sheet: the excel sheet with the diagnosis
    """
    # Get yolo prediction    
    yolo_pred, symptom_classes = detect_image(model = yolo_model, image_path = img_path, output_path = '', classes = './configurations/yolo_classes.txt', colab = False, score_threshold= 0.3, clahe = False)
    # Get glaucoma prediction and cdr
    glaucoma_pred, cdr_value = segment_image(crop_model = disc_model, seg_model = seg_model, DC_DATA_IMAGE = img_path)

    # Save mask for future reference
    glaucoma_pred = np.asarray(glaucoma_pred)
    im = Image.fromarray(glaucoma_pred)
    im.save('./uploads/masks/' + 'mask_' + file_name)

    # Reread pillow saved mask
    temp_img = cv2.imread('./uploads/masks/' + 'mask_' + file_name)
    temp_img = cv2.cvtColor(temp_img, cv2.COLOR_BGR2RGB)

    # Layer the glaucoma mask on top of the yolo base
    layer_images = overlay_image(base = yolo_pred, top = temp_img) #, cdr = cdr_value)
    im = Image.fromarray(layer_images)

    # Create spreadsheet of detailed diagnosis
    diagnosis_sheet = diagnose(symptom_classes = symptom_classes, cdr = cdr_value)
    # Save spreadsheet
    diagnosis_sheet.save('./uploads/sheets/' + 'summary_' + file_name[:-4] + '.xlsx')

    return im, diagnosis_sheet

def overlay_image(base, top):
    """
    Overlays the segmentation model on top of the optic-disc and optic-cup.

    Parameters:
        base: the base image, typically the yolodetection output
        top: top image, typically the segmentation image
    
    Returns:
        Overlayed image.
    """
    # Change all values to numpy array, just in case
    base = np.asarray(base)
    top = np.asarray(top)
    # Create a region of intrest
    rows,cols,channels = top.shape
    roi = base[0:rows, 0:cols]
    # Create mask of the mask and create inverse mask
    img2gray = cv2.cvtColor(top,cv2.COLOR_BGR2GRAY)
    ret, mask = cv2.threshold(img2gray, 10, 255, cv2.THRESH_BINARY)
    mask_inv = cv2.bitwise_not(mask)
    # Remove area of the base where the mask belongs to
    img1_bg = cv2.bitwise_and(roi,roi,mask = mask_inv)
    # Only use the part of the mask that is needed
    img2_fg = cv2.bitwise_and(top,top,mask = mask)
    img2_fg = np.asarray(img2_fg)
    # Combine the mask and the base
    dst = cv2.add(img1_bg, img2_fg)
    base[0:rows, 0:cols ] = dst

    x = base.shape[1]
    y = base.shape[0]

    return base

def diagnose(symptom_classes, cdr):
    """
    Diagnoses the Fundus Image based on the detections and segmentations.

    The Diagnosis:
        diabetic retinopathy: mild npdr: one MA, moderate npdr: hemorrhages, exudates, beading, severe: >20 hemorrhages, venous beading, IRMA, PDR: neovascularization, vitreous/preretinal hemorrhage
        myopia: severe myopia
        cataracts: appearance of cataract
        glaucoma: optic disc > 65 CDR
        age related macular degeneration: early armd: >20 drusen, intermediate armd: <20drusen, advanced armd: geography atrophy
    
    On the classes that CAN be detected, some are not used:
        0 = artery narrowing/ not used, 1 = artery nickening/ not used, 2 = cataract, 3 = choroidal nevus/ not used, 4 = copper wiring/ not used, 5 = cotton wool spots
        6 = drusen, 7 = edema, 8 = epiretinal membrance, 9 = georgraphy atrophy/ not used, 10 = glaucoma optic disk/ not used, 11 = hard exudates , 12 = healthy/ not used
        13 = healthy optic disc/ not used, 14 = hemorrhages, 15 = irma/ not used, 16 = microaneurysms , 17 = neovascularization/ not used, 18 = papilledema/ not used, 19 = preretinal hemorrhage/ not used
        20 = rpe depigmentation/ not used, 21 = refractive media opacity/ not used, 22 = venous beading/ not used, 23 = vitreous hemorrhage/ not used, 24 = myopia
        25 = chorioretinal atrophy/ not used, 26 = arterial wall thickening/ not used

    Parameters:
        symptom_classes: all of the classes that will be diagnosed.
        cdr: the cup to disc ratio

    Returns:
        Excel spreadsheet with the diagnosis.
    """

    symptom_classes = np.asarray(symptom_classes) # Double check if it is an np array
    diagnosis = [["Diagnosis", "LevelOrGrade", "Reasoning"]] # Create empty array for diagnosis
    diagnosis_xl = openpyxl.Workbook() # Create a blank spreadsheet
    diagnosis_sheet = diagnosis_xl.active
    diagnosis_sheet.title = "Diagnosis for E-EYE Prediction"

    # Diabetic retinopathy diagnosis
    if (symptom_classes[16] == 1 and symptom_classes[16] == 0 and symptom_classes[5] == 0 and symptom_classes[11] == 0 and symptom_classes[22] == 0 and symptom_classes[14] == 0): 
        diagnosis.append(["Diabetic Retinopathy", 'mild', 'ARDD detected ' + str(symptom_classes[16]) + " microaneurysms."])
    elif (symptom_classes[16] > 0 or symptom_classes[5] > 0 or symptom_classes[11] > 0 or symptom_classes[22] > 0 or (symptom_classes[14] > 0 and symptom_classes[14] < 20)): #one or more microaneurysms, cotton wool spots, hard exudates, hemorrhages, venous beading
        diagnosis.append(["Diabetic Retinopathy", 'severe', 'ARDD detected ' + str(symptom_classes[16]) + ' microaneurysms, ' + str(symptom_classes[5]) + ' cotton wool spots, ' + 
                            str(symptom_classes[11]) + ' hard exudates, ' + str(symptom_classes[22]) + ' cases of venous beading ' + ', and ' + str(symptom_classes[14]) + ' hemorrhages.'])
    elif (symptom_classes[16] > 0 or symptom_classes[5] > 0 or symptom_classes[11] > 0  or symptom_classes[15] > 0 or symptom_classes[22] > 0) and symptom_classes[14] > 20: #one or more microaneurysms, cotton wool spots, hard exudates >20, hemorrhages, venous beading, IRMA
        diagnosis.append(["Diabetic Retinopathy", 'severe', 'ARDD detected ' + str(symptom_classes[16]) + ' microaneurysms, ' + str(symptom_classes[5]) + ' cotton wool spots, ' + 
                            str(symptom_classes[11]) + ' hard exudates, ' + str(symptom_classes[15]) + ' cases of IRMA, '+ str(symptom_classes[22]) + ' cases of venous beading' + ', and ' + str(symptom_classes[14]) + ' hemorrhages.'])
    elif (symptom_classes[16] > 0 or symptom_classes[5] > 0 or symptom_classes[11] > 0 or symptom_classes[14] > 20 or symptom_classes[15] > 0 or symptom_classes[22] > 0) and (symptom_classes[17] > 0 or symptom_classes[19] > 0): #one or more microaneurysms, cotton wool spots, hard exudates >20, hemorrhages, venous beading, IRMA, neovascularization, vitreuos/prertinal hemorrhage
        diagnosis.append(["Diabetic Retinopathy", 'proliferative', 'ARDD detected ' + str(symptom_classes[16]) + ' microaneurysms, ' + str(symptom_classes[5]) + ' cotton wool spots, ' + 
                            str(symptom_classes[11]) + ' hard exudates, ' + str(symptom_classes[22]) + ' cases of venous beading' + ', and ' + str(symptom_classes[14]) + ' hemorrhages.'])
    
    '''
    ARDD no longer takes cataracts as a diagnosis.
    '''
    # Glaucoma Diagnosis
    if cdr > 0.650:
        diagnosis.append(["Glaucoma", 'not applicable', 'ARDD detected a vertical cup to disk ratio of ' + str(cdr) + '.'])

    # Pathological myopia diagnosis
    if symptom_classes[24] > 0:
        diagnosis.append(["Myopia", 'pathological', 'ARDD detected signs of pathological myopia.'])

    # AMD diagnosis
    if symptom_classes[6] > 0 and symptom_classes[6] < 20:
        diagnosis.append(["Age Related Macular Degeneration", 'early', 'ARDD detected ' + str(symptom_classes[6]) + " drusen."])
    elif symptom_classes[6] > 20:
        diagnosis.append(["Age Related Macular Degeneration", 'intermediate', 'ARDD detected ' + str(symptom_classes[6]) + " drusen."])   
    elif symptom_classes[6] > 20 or symptom_classes[9]:
        diagnosis.append(["Age Related Macular Degeneration", 'advanced', 'ARDD detected ' + str(symptom_classes[6]) + " drusen."]) 

    # Epiretinal membrane
    if symptom_classes[8] > 0:
        diagnosis.append(["Epiretinal Membrane", 'not applicable', 'ARDD detected the possibility of an Epiretinal Membrane' ]) 

    # Healthy
    if len(diagnosis) == 1:
        diagnosis.append(['None Found', "not applicable", "ARDD found no anomalies."])

    # Save predictions to workbook
    for diagnose in range(len(diagnosis)):
        disease = diagnosis[diagnose][0]
        stage = diagnosis[diagnose][1]
        reasons = diagnosis[diagnose][2]
        diagnosis_sheet.cell(row = diagnose + 1, column = 1).value = disease  
        diagnosis_sheet.cell(row = diagnose + 1, column = 2).value = stage  
        diagnosis_sheet.cell(row = diagnose + 1, column = 3).value = reasons  

    # Edit column dimensions
    diagnosis_sheet.column_dimensions['A'].width = 40
    diagnosis_sheet.column_dimensions['B'].width = 20
    diagnosis_sheet.column_dimensions['C'].width = 200

    # Edit column formats -> bold
    diagnosis_sheet.cell(row = 1, column = 1).font = openpyxl.styles.Font(bold = True)
    diagnosis_sheet.cell(row = 1, column = 2).font = openpyxl.styles.Font(bold = True)
    diagnosis_sheet.cell(row = 1, column = 3).font = openpyxl.styles.Font(bold = True)

    return diagnosis_xl

def base64_to_pil(img_base64):
    """
    Changes the image from base64 to pillow format.

    Parameters:
        img_base64: the image to be translated.

    Returns:
        pil_image: the image in the Pillow format.
    """
    image_data = re.sub('^data:image/.+;base64,', '', img_base64)
    pil_image = Image.open(BytesIO(base64.b64decode(image_data)))
    return pil_image


def np_to_base64(img_np):
    """
    Changes the image from numpy format to base64 format.

    Parameters:
        img_np: the image to be translated.
    
    Returns:
        The image in base64 format.
    """
    img = Image.fromarray(img_np.astype('uint8'), 'RGB')
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    return u"data:image/png;base64," + base64.b64encode(buffered.getvalue()).decode("ascii")
